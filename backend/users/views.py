from rest_framework import generics, permissions, status
from rest_framework.throttling import AnonRateThrottle
from rest_framework.response import Response
from rest_framework.views import APIView
from django.contrib.auth import authenticate, login, logout
from rest_framework.parsers import MultiPartParser, FormParser
from .serializers import UserSerializer, RegisterSerializer
from .models import PasswordResetRequest

class RegisterView(generics.CreateAPIView):
    serializer_class = RegisterSerializer
    permission_classes = [permissions.AllowAny]

class LoginView(APIView):
    permission_classes = [permissions.AllowAny]
    throttle_classes = [AnonRateThrottle]

    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')

        user = authenticate(username=username, password=password)

        if user:
            login(request, user)
            user_data = UserSerializer(user).data
            # Include first_login_required flag
            user_data['first_login_required'] = user.first_login_required
            return Response(user_data)
        return Response({'error': 'Invalid Credentials'}, status=status.HTTP_400_BAD_REQUEST)

class LogoutView(APIView):
    def post(self, request):
        logout(request)
        return Response({'message': 'Logged out successfully'})

from django.views.decorators.csrf import ensure_csrf_cookie
from django.utils.decorators import method_decorator

@method_decorator(ensure_csrf_cookie, name='dispatch')
class GetCSRFToken(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        return Response({'success': 'CSRF cookie set'})

class PublicStaffView(APIView):
    """Public endpoint to list staff with department info"""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        from academic.models import CourseOffering
        
        users = User.objects.filter(role__in=['DOCTOR', 'STAFF'])
        result = []
        
        for user in users:
            # Get department from first course offering
            dept_name = None
            offering = CourseOffering.objects.filter(doctor=user).select_related('level__department').first()
            if offering and offering.level and offering.level.department:
                dept_name = offering.level.department.name
            
            result.append({
                'id': user.id,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email,
                'role': user.role,
                'department': dept_name,
            })
        
        return Response(result)

class UserProfileView(generics.RetrieveUpdateAPIView):
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def get_object(self):
        return self.request.user

class ChangePasswordView(APIView):
    """
    Change password endpoint - particularly for first login requirement.
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')

        # Validate inputs
        if not current_password or not new_password or not confirm_password:
            return Response(
                {'error': 'All password fields are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check current password
        if not user.check_password(current_password):
            return Response(
                {'error': 'Current password is incorrect'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check new passwords match
        if new_password != confirm_password:
            return Response(
                {'error': 'New passwords do not match'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check new password is not the same as national_id
        if user.national_id and new_password == user.national_id:
            return Response(
                {'error': 'New password cannot be the same as your National ID'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check minimum password length
        if len(new_password) < 6:
            return Response(
                {'error': 'Password must be at least 6 characters'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update password
        user.set_password(new_password)
        user.first_login_required = False
        user.save()

        # Re-authenticate user with new password
        login(request, user)

        return Response({'message': 'Password changed successfully'})

from rest_framework import viewsets
from rest_framework.decorators import action
from .serializers import UserManagementSerializer
from django.contrib.auth import get_user_model
from .permissions import IsAdminRole
from .models import PasswordResetRequest

User = get_user_model()

class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserManagementSerializer
    permission_classes = [IsAdminRole]

    def perform_destroy(self, instance):
        try:
            instance.delete()
        except Exception as e:
            # Check for ProtectedError (cannot import easily inside method without overhead, so strict checking is tricky, 
            # but usually it's raised as ProtectedError)
            if 'ProtectedError' in str(type(e)):
                 from rest_framework.exceptions import ValidationError
                 raise ValidationError({'error': 'لا يمكن حذف هذا المستخدم لأنه مرتبط ببيانات أخرى (مثل جداول محاضرات أو درجات).'})
            
            from rest_framework.exceptions import APIException
            raise APIException(f"Deletion failed: {str(e)}")

    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        """Reset user password to their national ID"""
        user = self.get_object()
        if hasattr(user, 'national_id') and user.national_id:
            user.set_password(user.national_id)
            user.first_login_required = True
            user.save()
            return Response({'message': 'تم إعادة تعيين كلمة المرور بنجاح'})
        return Response({'error': 'لا يمكن إعادة تعيين كلمة المرور - الرقم القومي غير موجود'}, 
                        status=status.HTTP_400_BAD_REQUEST)

class PasswordResetRequestView(APIView):
    """View to allow users to request a password reset"""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        """List requests made by this user"""
        if not request.user.is_authenticated:
            return Response({'error': 'Authentication required'}, status=status.HTTP_401_UNAUTHORIZED)
            
        requests = PasswordResetRequest.objects.filter(user=request.user).order_by('-requested_at')
        result = []
        for req in requests:
            result.append({
                'id': req.id,
                'reason': req.reason,
                'status': req.status,
                'status_display': req.get_status_display(),
                'requested_at': req.requested_at,
            })
        return Response(result)

    def post(self, request):
        """Create a new password reset request"""
        if request.user.is_authenticated:
            user = request.user
        else:
            national_id = request.data.get('national_id')
            if not national_id:
                return Response({'error': 'National ID is required'}, status=status.HTTP_400_BAD_REQUEST)
            try:
                user = User.objects.get(national_id=national_id)
            except User.DoesNotExist:
                return Response({'error': 'User with this National ID not found'}, status=status.HTTP_404_NOT_FOUND)

        # Check if already pending
        if PasswordResetRequest.objects.filter(user=user, status=PasswordResetRequest.Status.PENDING).exists():
            return Response(
                {'error': 'لديك طلب قيد الانتظار بالفعل'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
            
        reason = request.data.get('reason', '')
        req = PasswordResetRequest.objects.create(user=user, reason=reason)
        
        return Response({
            'message': 'تم إرسال طلبك بنجاح',
            'id': req.id
        }, status=status.HTTP_201_CREATED)


class AdminPasswordResetRequestsView(APIView):
    """Admin view for approving/rejecting password reset requests"""
    permission_classes = [IsAdminRole]

    def get(self, request):
        """List all pending requests"""
        from django.utils import timezone
        
        status_filter = request.query_params.get('status', 'PENDING')
        requests = PasswordResetRequest.objects.filter(status=status_filter).select_related('user')
        
        result = []
        for req in requests:
            result.append({
                'id': req.id,
                'user_id': req.user.id,
                'user_name': f"{req.user.first_name} {req.user.last_name}".strip(),
                'user_role': req.user.get_role_display(),
                'national_id': req.user.national_id,
                'reason': req.reason,
                'status': req.status,
                'status_display': req.get_status_display(),
                'requested_at': req.requested_at,
            })
        return Response(result)

    def post(self, request, pk):
        """Approve or reject request"""
        from django.utils import timezone
        
        try:
            req = PasswordResetRequest.objects.get(id=pk)
            action = request.data.get('action')
            
            if action == 'approve':
                # Reset password to national ID
                user = req.user
                if user.national_id:
                    user.set_password(user.national_id)
                    user.first_login_required = True
                    user.save()
                    
                    req.status = PasswordResetRequest.Status.APPROVED
                    req.reviewed_at = timezone.now()
                    req.reviewed_by = request.user
                    req.save()
                    
                    return Response({'message': 'تم الموافقة على الطلب وإعادة تعيين كلمة المرور'})
                else:
                    return Response(
                        {'error': 'لا يوجد رقم قومي للمستخدم، لا يمكن إعادة تعيين كلمة المرور'}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )
                    
            elif action == 'reject':
                req.status = PasswordResetRequest.Status.REJECTED
                req.reviewed_at = timezone.now()
                req.reviewed_by = request.user
                req.save()
                return Response({'message': 'تم رفض الطلب'})
                
            else:
                return Response({'error': 'Invalid action'}, status=status.HTTP_400_BAD_REQUEST)
                
        except PasswordResetRequest.DoesNotExist:
            return Response({'error': 'Request not found'}, status=status.HTTP_404_NOT_FOUND)


class UploadUsersView(APIView):
    """Admin endpoint to bulk upload users via Excel file"""
    permission_classes = [IsAdminRole]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No file provided'}, status=status.HTTP_400_BAD_REQUEST)

        # Check file extension
        if not file.name.endswith(('.xlsx', '.xls', '.csv')):
            return Response(
                {'error': 'Unsupported file format. Please upload an Excel (.xlsx, .xls) or CSV file.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            import openpyxl
            import csv
            import io

            users_created = []
            errors = []

            if file.name.endswith('.csv'):
                decoded = file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

            for i, row in enumerate(rows, start=2):
                try:
                    username = str(row.get('username', '') or '').strip()
                    national_id = str(row.get('national_id', '') or '').strip()
                    first_name = str(row.get('first_name', '') or '').strip()
                    last_name = str(row.get('last_name', '') or '').strip()
                    email = str(row.get('email', '') or '').strip()
                    role = str(row.get('role', 'STUDENT') or 'STUDENT').strip().upper()

                    if not username:
                        errors.append(f'Row {i}: username is required')
                        continue

                    if User.objects.filter(username=username).exists():
                        errors.append(f'Row {i}: username "{username}" already exists')
                        continue

                    if national_id and User.objects.filter(national_id=national_id).exists():
                        errors.append(f'Row {i}: national_id "{national_id}" already exists')
                        continue

                    # Validate role
                    valid_roles = [choice[0] for choice in User.Role.choices]
                    if role not in valid_roles:
                        errors.append(f'Row {i}: invalid role "{role}". Valid roles: {", ".join(valid_roles)}')
                        continue

                    user = User(
                        username=username,
                        national_id=national_id or None,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        role=role,
                    )
                    # Password defaults to national_id via the model's save() method
                    user.save()
                    users_created.append(username)

                except Exception as e:
                    errors.append(f'Row {i}: {str(e)}')

            return Response({
                'message': f'Successfully created {len(users_created)} users',
                'created_count': len(users_created),
                'users_created': users_created,
                'errors': errors,
                'error_count': len(errors),
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

