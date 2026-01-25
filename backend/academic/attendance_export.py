"""
Attendance Excel Export Utility

Automatically exports attendance records to Excel files.
"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.conf import settings


def export_attendance_to_excel(course_offering, attendance_date, attendance_records):
    """
    Export attendance records to an Excel file.
    
    Args:
        course_offering: CourseOffering instance
        attendance_date: Date of attendance
        attendance_records: List of Attendance objects
        
    Returns:
        str: Path to the generated Excel file
    """
    # Create media/attendance directory if it doesn't exist
    attendance_dir = os.path.join(settings.MEDIA_ROOT, 'attendance')
    os.makedirs(attendance_dir, exist_ok=True)
    
    # Generate filename
    date_str = attendance_date.strftime('%Y-%m-%d')
    subject_code = course_offering.subject.code if course_offering.subject else 'UNKNOWN'
    level_name = course_offering.level.get_name_display() if course_offering.level else 'UNKNOWN'
    filename = f"attendance_{subject_code}_{level_name}_{date_str}.xlsx"
    filepath = os.path.join(attendance_dir, filename)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Header information
    ws.merge_cells('A1:E1')
    ws['A1'] = f"سجل الحضور - {course_offering.subject.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"التاريخ: {date_str}"
    ws['A2'].font = Font(size=11)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:E3')
    doctor_name = f"{course_offering.doctor.first_name} {course_offering.doctor.last_name}".strip() or course_offering.doctor.username if course_offering.doctor else "غير محدد"
    ws['A3'] = f"المحاضر: {doctor_name}"
    ws['A3'].font = Font(size=11)
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = ['#', 'الرقم القومي', 'اسم الطالب', 'الحالة', 'ملاحظات']
    header_row = 5
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row_num = header_row + 1
    for idx, attendance in enumerate(attendance_records, 1):
        ws.cell(row=row_num, column=1, value=idx).border = border
        ws.cell(row=row_num, column=2, value=attendance.student.national_id).border = border
        ws.cell(row=row_num, column=3, value=attendance.student.full_name).border = border
        
        # Status with color coding
        status_cell = ws.cell(row=row_num, column=4, value=attendance.get_status_display())
        status_cell.border = border
        status_cell.alignment = Alignment(horizontal='center')
        
        if attendance.status == 'PRESENT':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100")
        elif attendance.status == 'ABSENT':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006")
        else:  # EXCUSED
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C6500")
        
        ws.cell(row=row_num, column=5, value="").border = border  # Notes column
        row_num += 1
    
    # Summary statistics
    summary_row = row_num + 2
    total_students = len(attendance_records)
    present_count = sum(1 for a in attendance_records if a.status == 'PRESENT')
    absent_count = sum(1 for a in attendance_records if a.status == 'ABSENT')
    excused_count = sum(1 for a in attendance_records if a.status == 'EXCUSED')
    
    ws.cell(row=summary_row, column=1, value="إحصائيات الحضور:").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"إجمالي الطلاب: {total_students}")
    ws.cell(row=summary_row + 2, column=1, value=f"حاضر: {present_count}")
    ws.cell(row=summary_row + 2, column=1).font = Font(color="006100")
    ws.cell(row=summary_row + 3, column=1, value=f"غائب: {absent_count}")
    ws.cell(row=summary_row + 3, column=1).font = Font(color="9C0006")
    ws.cell(row=summary_row + 4, column=1, value=f"بعذر: {excused_count}")
    ws.cell(row=summary_row + 4, column=1).font = Font(color="9C6500")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    
    # Save file
    wb.save(filepath)
    
    # Return relative path for URL generation
    return os.path.join('attendance', filename)
